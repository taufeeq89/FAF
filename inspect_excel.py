from openpyxl import load_workbook
path = r'C:\Users\tauft\PROJECTS\FAF\ZakatCalculator2026.xlsx'
wb = load_workbook(path, data_only=True)
print(wb.sheetnames)
for ws in wb.worksheets:
    print('---', ws.title, '---')
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        print(row)
    print()
